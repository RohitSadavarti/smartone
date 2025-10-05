import csv
import os
from io import StringIO
from flask import Flask, Response, jsonify, request, session, redirect, url_for, render_template, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename
import logging
from datetime import datetime, timedelta
from openpyxl import load_workbook
import psycopg2
from functools import wraps
from werkzeug.security import generate_password_hash
import numpy as np
import pandas as pd
import seaborn as sns
import matplotlib
matplotlib.use('Agg') # Use non-interactive backend for web servers
import matplotlib.pyplot as plt
import matplotlib.patches as patches # Import patches for gauge chart
from sqlalchemy import create_engine
import io

app = Flask(__name__)
CORS(app)

app.secret_key = 'your-secret-key-change-this-in-production'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SECURE'] = False
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)
app.config['UPLOAD_FOLDER'] = 'uploads/'
app.config['ALLOWED_EXTENSIONS'] = {'xlsx', 'xls', 'csv'}
app.config['PG_HOST'] = 'aws-0-ap-south-1.pooler.supabase.com'
app.config['PG_USER'] = 'postgres.avqpzwgdylnklbkyqukp'
app.config['PG_PASSWORD'] = 'asBjLmDfKfoZPVt9'
app.config['PG_DB'] = 'postgres'
app.config['sslmode']='require'

logging.basicConfig(level=logging.INFO)

# --- Database Connection for Dashboard---
DATABASE_URL = f"postgresql+psycopg2://{app.config['PG_USER']}:{app.config['PG_PASSWORD']}@{app.config['PG_HOST']}/{app.config['PG_DB']}?sslmode=require"
engine = create_engine(DATABASE_URL)

_df_cache = None
def get_data():
    """Fetches and caches data from the database."""
    global _df_cache
    if _df_cache is None:
        try:
            df = pd.read_sql("SELECT * FROM attendance;", engine)
            if not df.empty:
                df["date"] = pd.to_datetime(df["date"])
                df["day"] = df["date"].dt.day_name()
            _df_cache = df
        except Exception as e:
            print(f"Error loading data: {e}")
            return pd.DataFrame()
    return _df_cache

def get_pg_connection():
    return psycopg2.connect(
        host=app.config['PG_HOST'],
        database=app.config['PG_DB'],
        user=app.config['PG_USER'],
        password=app.config['PG_PASSWORD'],
        port=6543,
        sslmode=app.config['sslmode']
    )

def get_user_from_db(email):
    try:
        connection = get_pg_connection()
        cursor = connection.cursor()
        cursor.execute("""
            SELECT id, email, password_hash, created_at, role
            FROM users
            WHERE email = %s
        """, (email,))
        user_row = cursor.fetchone()
        cursor.close()
        connection.close()
        if user_row:
            return {
                'id': user_row[0],
                'email': user_row[1],
                'password_hash': user_row[2],
                'created_at': user_row[3],
                'role': user_row[4] if user_row[4] else 'user',
                'name': user_row[1].split('@')[0].replace('.', ' ').title(),
            }
        return None
    except Exception as e:
        app.logger.error(f"Database error in get_user_from_db: {e}")
        return None

def verify_password_crypt(password, password_hash):
    try:
        connection = get_pg_connection()
        cursor = connection.cursor()
        cursor.execute("""
            SELECT crypt(%s, %s) = %s AS password_match
        """, (password, password_hash, password_hash))
        result = cursor.fetchone()
        cursor.close()
        connection.close()
        return result[0] if result else False
    except Exception as e:
        app.logger.error(f"Password verification error: {e}")
        return False

def create_user_with_crypt(email, password, role='user'):
    try:
        connection = get_pg_connection()
        cursor = connection.cursor()
        cursor.execute("""
            INSERT INTO users (email, password_hash, role)
            VALUES (%s, crypt(%s, gen_salt('bf')), %s)
            ON CONFLICT (email) DO UPDATE
            SET password_hash = crypt(EXCLUDED.password_hash, gen_salt('bf')),
                role = EXCLUDED.role
        """, (email, password, role))
        connection.commit()
        cursor.close()
        connection.close()
        return True
    except Exception as e:
        app.logger.error(f"Error creating user: {e}")
        return False

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.is_json or request.headers.get('Content-Type') == 'application/json' or 'api/' in request.path:
                return jsonify({'error': 'Authentication required', 'redirect': '/login'}), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.is_json or request.headers.get('Content-Type') == 'application/json' or 'api/' in request.path:
                return jsonify({'error': 'Authentication required', 'redirect': '/login'}), 401
            return redirect(url_for('login_page'))

        user_role = session.get('user_role', 'user')
        if user_role != 'admin':
            if request.is_json or request.headers.get('Content-Type') == 'application/json' or 'api/' in request.path:
                return jsonify({
                    'error': 'Admin access required',
                    'message': 'You do not have permission to access this resource',
                    'redirect': '/'
                }), 403
            return redirect(url_for('home'))

        return f(*args, **kwargs)
    return decorated_function

@app.route('/login')
def login_page():
    if 'user_id' in session:
        return redirect(url_for('home'))
    return render_template('login.html')

@app.route("/")
def root():
    if 'user_id' not in session:
        return redirect(url_for('login_page'))
    return redirect(url_for('home'))

@app.route("/home")
@login_required
def home():
    return render_template("index.html")

@app.route("/attendance")
@login_required
def attendance_tracker():
    return render_template("Attendance.html")

@app.route("/admin-students")
@admin_required
def admin_student():
    return render_template("admin-students.html")

@app.route("/admin-teacher")
@admin_required
def admin_teacher():
    return render_template("admin-teacher.html")

@app.route("/dashboard")
@login_required
def dashboard_page():
    return render_template("dashboard.html")

@app.route('/static/<path:filename>')
def serve_static(filename):
    return send_from_directory('static', filename)

@app.route('/api/login', methods=['POST'])
def api_login():
    try:
        data = request.get_json()
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        remember = data.get('remember', False)
        user = get_user_from_db(email)

        if user and verify_password_crypt(password, user['password_hash']):
            session.clear()
            session['user_id'] = user['email']
            session['user_name'] = user['name']
            session['user_role'] = user['role']
            session['user_db_id'] = user['id']
            session.permanent = remember

            return jsonify({
                'success': True,
                'message': 'Login successful',
                'redirect': '/home',
                'user': {
                    'email': user['email'],
                    'name': user['name'],
                    'role': user['role']
                }
            })
        else:
            return jsonify({'success': False, 'message': 'Invalid email or password'}), 401
    except Exception as e:
        return jsonify({'success': False, 'message': 'Login failed. Please try again.'}), 500

@app.route('/api/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({'success': True, 'message': 'Logged out successfully', 'redirect': '/login'})

@app.route('/api/forgot-password', methods=['POST'])
def api_forgot_password():
    try:
        data = request.get_json()
        email = data.get('email', '').strip().lower()
        if not email:
            return jsonify({'success': False, 'message': 'Email is required'}), 400
        user = get_user_from_db(email)
        if not user:
            return jsonify({'success': False, 'message': 'Email not found'}), 404
        return jsonify({'success': True, 'message': 'Password reset link sent to your email'})
    except Exception as e:
        return jsonify({'success': False, 'message': 'Failed to send reset link'}), 500

@app.route('/api/register', methods=['POST'])
def api_register():
    try:
        data = request.get_json()
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        role = data.get('role', 'user')
        if role not in ['admin', 'user']:
            role = 'user'
        if not email or not password:
            return jsonify({'success': False, 'message': 'Email and password are required'}), 400
        if len(password) < 6:
            return jsonify({'success': False, 'message': 'Password must be at least 6 characters long'}), 400
        existing_user = get_user_from_db(email)
        if existing_user:
            return jsonify({'success': False, 'message': 'User already exists'}), 409
        if create_user_with_crypt(email, password, role):
            return jsonify({'success': True, 'message': f'User {email} created successfully with role {role}'})
        else:
            return jsonify({'success': False, 'message': 'Registration failed. Please try again.'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': 'Registration failed. Please try again.'}), 500

@app.route('/api/user-info', methods=['GET'])
@login_required
def api_user_info():
    return jsonify({
        'user_id': session.get('user_id'),
        'user_name': session.get('user_name'),
        'user_role': session.get('user_role'),
        'user_db_id': session.get('user_db_id')
    })

@app.route('/departments', methods=['GET'])
@login_required
def get_departments():
    connection = get_pg_connection()
    cursor = connection.cursor()
    cursor.execute("SELECT DISTINCT department FROM Teachers ORDER BY department ASC")
    departments = cursor.fetchall()
    cursor.close()
    connection.close()
    return jsonify([dept[0] for dept in departments])

@app.route('/teachers', methods=['GET'])
@login_required
def get_teachers():
    department = request.args.get('department')
    connection = get_pg_connection()
    cursor = connection.cursor()
    cursor.execute("SELECT DISTINCT teacher_name FROM Teachers WHERE department = %s ORDER BY teacher_name ASC", (department,))
    teachers = cursor.fetchall()
    cursor.close()
    connection.close()
    return jsonify([teacher[0] for teacher in teachers])

@app.route('/student-classes', methods=['GET'])
@login_required
def get_student_classes():
    connection = get_pg_connection()
    cursor = connection.cursor()
    cursor.execute("SELECT DISTINCT class FROM Students ORDER BY class ASC")
    classes = cursor.fetchall()
    cursor.close()
    connection.close()
    return jsonify([class_[0] for class_ in classes])

@app.route('/subjects', methods=['GET'])
@login_required
def get_subjects():
    department = request.args.get('department')
    class_name = request.args.get('class')
    teacher_name = request.args.get('teacher_name')
    if not all([department, class_name, teacher_name]):
        return jsonify({"error": "Parameters 'department', 'class', and 'teacher_name' are required"}), 400
    connection = get_pg_connection()
    cursor = connection.cursor()
    query = "SELECT DISTINCT subject FROM Teachers WHERE department = %s AND class = %s AND teacher_name = %s ORDER BY subject ASC"
    cursor.execute(query, (department, class_name, teacher_name))
    subjects = cursor.fetchall()
    cursor.close()
    connection.close()
    return jsonify([subject[0] for subject in subjects])

@app.route('/time-slots', methods=['GET'])
@login_required
def get_time_slots():
    connection = get_pg_connection()
    cursor = connection.cursor()

    cursor.execute("SELECT DISTINCT time_slot FROM Teachers WHERE time_slot IS NOT NULL AND time_slot != ''")
    time_slots_raw = cursor.fetchall()
    cursor.close()
    connection.close()

    time_slots_list = [item[0] for item in time_slots_raw]

    def sort_key(time_slot_str):
        try:
            start_time_str = time_slot_str.split('-')[0].strip()
            hour, minute = map(int, start_time_str.split(':'))

            if hour < 8:
                hour += 12

            return (hour, minute)

        except (ValueError, IndexError):
            logging.warning(f"Could not parse time slot: '{time_slot_str}'. Placing it at the end.")
            return (23, 59)

    sorted_time_slots = sorted(time_slots_list, key=sort_key)

    return jsonify(sorted_time_slots)

@app.route('/students', methods=['GET'])
@login_required
def get_students():
    query_param = request.args.get('query', '').strip()
    department = request.args.get('department', '').strip()
    student_class = request.args.get('class', '').strip()
    connection = get_pg_connection()
    cursor = connection.cursor()
    if query_param:
        cursor.execute("SELECT roll_number, name FROM Students WHERE roll_number LIKE %s OR name LIKE %s", (f"%{query_param}%", f"%{query_param}%"))
    elif department and student_class:
        cursor.execute("SELECT roll_number, name FROM Students WHERE department = %s AND class = %s", (department, student_class))
    elif department:
        cursor.execute("SELECT roll_number, name FROM Students WHERE department = %s", (department,))
    elif student_class:
        cursor.execute("SELECT roll_number, name FROM Students WHERE class = %s", (student_class,))
    else:
        cursor.execute("SELECT roll_number, name FROM Students")
    students = cursor.fetchall()
    cursor.close()
    connection.close()
    return jsonify([{"roll_number": student[0], "name": student[1]} for student in students])

@app.route('/attendance', methods=['POST'])
@login_required
def save_attendance():
    data = request.json
    connection = get_pg_connection()
    cursor = connection.cursor()

    for record in data['attendance_records']:
        cursor.execute("""
            SELECT id FROM Attendance
            WHERE date = %s AND lecture_time = %s AND roll_number = %s
        """, (record['date'], record['lecture_time'], record['roll_number']))

        existing = cursor.fetchone()

        if existing:
            cursor.execute("""
                UPDATE Attendance
                SET attendance = %s,
                    name = %s,
                    department = %s,
                    class = %s,
                    subject = %s,
                    teacher_name = %s
                WHERE date = %s AND lecture_time = %s AND roll_number = %s
            """, (
                record['attendance'],
                record['name'],
                record['department'],
                record['class'],
                record['subject'],
                record['teacher_name'],
                record['date'],
                record['lecture_time'],
                record['roll_number']
            ))
        else:
            cursor.execute("""
                INSERT INTO Attendance (date, roll_number, name, department, class, subject, teacher_name, lecture_time, attendance)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                record['date'],
                record['roll_number'],
                record['name'],
                record['department'],
                record['class'],
                record['subject'],
                record['teacher_name'],
                record['lecture_time'],
                record['attendance']
            ))

    connection.commit()
    cursor.close()
    connection.close()
    return jsonify({"status": "success"})

@app.route('/attendance-data', methods=['GET'])
@login_required
def get_attendance_data():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    department = request.args.get('department', '').strip()
    class_name = request.args.get('class', '').strip()
    if not start_date or not end_date:
        return jsonify({'message': 'Start date and end date are required'}), 400
    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'message': 'Invalid date format. Use YYYY-MM-DD.'}), 400
    try:
        connection = get_pg_connection()
        if connection is None:
            return jsonify({'message': 'Database connection failed'}), 500
        with connection.cursor() as cursor:
            query = "SELECT date, roll_number, name, department, class, attendance, lecture_time FROM Attendance WHERE date BETWEEN %s AND %s"
            filters = [start_date, end_date]
            if department:
                query += " AND department = %s"
                filters.append(department)
            if class_name:
                query += " AND class = %s"
                filters.append(class_name)
            cursor.execute(query, tuple(filters))
            rows = cursor.fetchall()
            if not rows:
                return jsonify([])
            result = [{'date': row[0], 'roll_number': row[1], 'name': row[2], 'department': row[3], 'class': row[4], 'attendance': row[5], 'lecture_time': row[6]} for row in rows]
            connection.close()
            return jsonify(result)
    except Exception as e:
        logging.error(f"Error fetching attendance data: {e}")
        return jsonify({'message': 'An error occurred while fetching attendance data'}), 500

@app.route('/attendance-csv', methods=['GET'])
@login_required
def download_attendance_csv():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    if not start_date or not end_date:
        return jsonify({'message': 'Start and end dates are required'}), 400
    try:
        connection = get_pg_connection()
        cursor = connection.cursor()
        query = "SELECT date, roll_number, name, department, class, attendance, lecture_time FROM Attendance WHERE date BETWEEN %s AND %s"
        cursor.execute(query, (start_date, end_date))
        rows = cursor.fetchall()
        cursor.close()
        connection.close()
        if not rows:
            return jsonify({'message': 'No attendance records found for the given dates'}), 404
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['Date', 'Roll Number', 'Name', 'Department', 'Class', 'Attendance', 'Lecture Time'])
        writer.writerows(rows)
        output.seek(0)
        return Response(output.getvalue(), mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=attendance_data.csv'})
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'message': 'An error occurred while generating the CSV'}), 500

@app.route('/student-departments', methods=['GET'])
@login_required
def get_student_departments():
    connection = get_pg_connection()
    cursor = connection.cursor()
    cursor.execute("SELECT DISTINCT department FROM Students ORDER BY department ASC")
    departments = cursor.fetchall()
    cursor.close()
    connection.close()
    return jsonify([dept[0] for dept in departments])

@app.route('/submit', methods=['POST'])
@admin_required
def submit_student():
    try:
        data = request.json
        roll_number = data.get('roll_number')
        name = data.get('name')
        department = data.get('department')
        class_value = data.get('class')
        if not all([roll_number, name, department, class_value]):
            return jsonify({'success': False, 'message': 'All fields are required'}), 400
        connection = get_pg_connection()
        cursor = connection.cursor()
        cursor.execute("SELECT 1 FROM Students WHERE roll_number = %s", (roll_number,))
        existing_student = cursor.fetchone()
        if existing_student:
            cursor.close()
            connection.close()
            return jsonify({'success': False, 'message': 'Roll number already exists'}), 409
        cursor.execute("""
            INSERT INTO Students (roll_number, name, department, class)
            VALUES (%s, %s, %s, %s)
        """, (roll_number, name, department, class_value))
        connection.commit()
        cursor.close()
        connection.close()
        return jsonify({'success': True, 'message': 'Student added successfully'})
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'success': False, 'message': 'Failed to submit student data'}), 500

def process_excel(file_path):
    data = []
    workbook = load_workbook(file_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return data

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

@app.route('/upload-file', methods=['POST'])
@admin_required
def upload_file():
    try:
        file = request.files.get('file')
        if not file or file.filename == '':
            return jsonify({'message': 'No file selected'}), 400
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            file.save(file_path)
            data = process_excel(file_path)
            total_records = len(data)
            valid_records = 0
            error_records = 0
            valid_data = []
            error_data = []
            connection = get_pg_connection()
            cursor = connection.cursor()
            for row in data:
                try:
                    roll_number, name, department, class_value = row
                    cursor.execute("SELECT 1 FROM Students WHERE roll_number = %s", (roll_number,))
                    existing = cursor.fetchone()
                    if existing:
                        error_data.append((roll_number, name, department, class_value, "Duplicate Roll Number"))
                        error_records += 1
                        continue
                    cursor.execute("""
                        INSERT INTO Students (roll_number, name, department, class)
                        VALUES (%s, %s, %s, %s)
                    """, (roll_number, name, department, class_value))
                    valid_data.append((roll_number, name, department, class_value))
                    valid_records += 1
                except Exception as e:
                    error_data.append(row + (str(e),))
                    error_records += 1
            if valid_data:
                cursor.executemany("INSERT INTO Validrecords (roll_number, name, department, class) VALUES (%s, %s, %s, %s)", valid_data)
            if error_data:
                cursor.executemany("INSERT INTO Errorrecords (roll_number, name, department, class, error_message) VALUES (%s, %s, %s, %s, %s)", error_data)
            upload_time = datetime.now()
            cursor.execute("INSERT INTO UploadHistory (uploader_name, total_records, valid_records, error_records, upload_time) VALUES (%s, %s, %s, %s, %s)", (session.get('user_name', 'Unknown'), total_records, valid_records, error_records, upload_time))
            connection.commit()
            cursor.close()
            connection.close()
            return jsonify({'message': 'File processed successfully.', 'total_records': total_records, 'valid_records': valid_records, 'error_records': error_records}), 200
        else:
            return jsonify({'message': 'Invalid file type.'}), 400
    except Exception as e:
        print(f"Error while uploading file: {e}")
        return jsonify({'message': 'An error occurred while processing the file.'}), 500

@app.route('/upload-history', methods=['GET'])
@admin_required
def upload_history():
    try:
        connection = get_pg_connection()
        cursor = connection.cursor()
        cursor.execute("SELECT uploader_name, total_records, valid_records, error_records, upload_time FROM UploadHistory")
        rows = cursor.fetchall()
        history = [{'uploader_name': row[0], 'total_records': row[1], 'valid_records': row[2], 'error_records': row[3], 'upload_time': row[4]} for row in rows]
        cursor.close()
        connection.close()
        return jsonify(history)
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'message': 'Failed to fetch upload history'}), 500

@app.route('/valid-records-csv', methods=['GET'])
@admin_required
def download_valid_csv():
    try:
        connection = get_pg_connection()
        cursor = connection.cursor()
        cursor.execute("SELECT roll_number, name, department, class FROM Validrecords")
        rows = cursor.fetchall()
        cursor.close()
        connection.close()
        if not rows:
            return jsonify({'message': 'No valid records found'}), 404
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['Date', 'Roll Number', 'Name', 'Department', 'Class', 'Attendance', 'Lecture Time'])
        writer.writerows(rows)
        output.seek(0)
        return Response(output.getvalue(), mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=valid_records.csv'})
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'message': 'An error occurred while generating the CSV'}), 500

@app.route('/upload-history/row-count', methods=['GET'])
@admin_required
def get_row_count():
    try:
        uploader_name = request.args.get('uploader_name', '').strip()
        if not uploader_name:
            return jsonify({'message': 'Uploader name is required'}), 400
        connection = get_pg_connection()
        cursor = connection.cursor()
        cursor.execute("SELECT COUNT(*) FROM UploadHistory WHERE uploader_name = %s", (uploader_name,))
        row_count = cursor.fetchone()[0]
        cursor.close()
        connection.close()
        return jsonify({'uploader_name': uploader_name, 'row_count': row_count})
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'message': 'An error occurred while fetching row count'}), 500

@app.route('/teachers', methods=['POST'])
@admin_required
def save_teacher():
    try:
        data = request.json
        teacher_name = data.get('teacher_name')
        day = data.get('day')
        subject = data.get('subject')
        time_slot = data.get('time_slot')
        department = data.get('department')
        class_value = data.get('class')
        if not all([teacher_name, day, subject, time_slot, department, class_value]):
            return jsonify({'message': 'All fields are required'}), 400
        connection = get_pg_connection()
        cursor = connection.cursor()
        query = "INSERT INTO Teachers (teacher_name, day, subject, time_slot, department, class) VALUES (%s, %s, %s, %s, %s, %s) ON CONFLICT (subject, time_slot, department, class) DO UPDATE SET teacher_name = EXCLUDED.teacher_name"
        cursor.execute(query, (teacher_name, day, subject, time_slot, department, class_value))
        connection.commit()
        cursor.close()
        connection.close()
        return jsonify({'message': 'Teacher data saved or updated successfully'})
    except Exception as e:
        print(f"Error saving teacher: {e}")
        return jsonify({'message': 'Failed to save teacher data'}), 500

@app.route('/upload-teachers', methods=['POST'])
@admin_required
def upload_teachers():
    try:
        file = request.files.get('file')
        if not file or file.filename == '':
            return jsonify({'message': 'No file selected'}), 400
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            file.save(file_path)
            data = process_excel(file_path)
            total_records = len(data)
            valid_records = 0
            error_records = 0
            valid_data = []
            error_data = []
            connection = get_pg_connection()
            cursor = connection.cursor()
            for row in data:
                try:
                    teacher_name, day, subject, time_slot, department, class_value = row
                    cursor.execute("INSERT INTO Teachers (teacher_name, day, subject, time_slot, department, class) VALUES (%s, %s, %s, %s, %s, %s) ON CONFLICT (subject, time_slot, department, class) DO UPDATE SET teacher_name = EXCLUDED.teacher_name", (teacher_name, day, subject, time_slot, department, class_value))
                    valid_data.append((teacher_name, day, subject, time_slot, department, class_value))
                    valid_records += 1
                except Exception as e:
                    error_data.append(row + (str(e),))
                    error_records += 1
            connection.commit()
            if valid_data:
                cursor.executemany("INSERT INTO ValidTeachers (teacher_name, day, subject, time_slot, department, class) VALUES (%s, %s, %s, %s, %s, %s)", valid_data)
            if error_data:
                cursor.executemany("INSERT INTO ErrorTeachers (teacher_name, day, subject, time_slot, department, class, error_message) VALUES (%s, %s, %s, %s, %s, %s, %s)", error_data)
            connection.commit()
            cursor.close()
            connection.close()
            return jsonify({'message': 'File processed successfully.', 'total_records': total_records, 'valid_records': valid_records, 'error_records': error_records}), 200
        else:
            return jsonify({'message': 'Invalid file type. Only .xlsx, .xls, and .csv files are allowed.'}), 400
    except Exception as e:
        print(f"Error while uploading file: {e}")
        return jsonify({'message': 'An error occurred while processing the file.'}), 500

# Admin utility endpoint to create users - ADMIN ONLY
@app.route('/api/admin/create-user', methods=['POST'])
@admin_required
def admin_create_user():
    try:
        data = request.get_json()
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        role = data.get('role', 'user')
        if role not in ['admin', 'user']:
            role = 'user'
        if not email or not password:
            return jsonify({'success': False, 'message': 'Email and password are required'}), 400
        if create_user_with_crypt(email, password, role):
            return jsonify({'success': True, 'message': f'User {email} created successfully with role {role}'})
        else:
            return jsonify({'success': False, 'message': 'Failed to create user'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': 'Failed to create user'}), 500

@app.route('/api/user-permissions', methods=['GET'])
@login_required
def api_user_permissions():
    user_role = session.get('user_role', 'user')
    permissions = {
        'can_access_admin_students': user_role == 'admin',
        'can_access_admin_teachers': user_role == 'admin',
        'can_upload_files': user_role == 'admin',
        'can_manage_students': user_role == 'admin',
        'can_manage_teachers': user_role == 'admin',
        'can_view_upload_history': user_role == 'admin',
        'user_role': user_role
    }
    return jsonify(permissions)

@app.route('/debug/session')
def debug_session():
    return jsonify({'session_data': dict(session), 'user_id_in_session': 'user_id' in session, 'session_keys': list(session.keys()), 'session_permanent': session.permanent})

@app.route('/test/session')
def test_session():
    session['test'] = 'working'
    return jsonify({'message': 'Session test set'})

@app.route('/test/session/check')
def check_session():
    return jsonify({'test_value': session.get('test', 'not found')})

@app.route('/health', methods=['GET'])
def health_check():
    try:
        # Test database connection
        connection = get_pg_connection()
        cursor = connection.cursor()
        cursor.execute("SELECT 1")
        cursor.close()
        connection.close()
        return jsonify({'status': 'healthy', 'database': 'connected', 'timestamp': datetime.now().isoformat()})
    except Exception as e:
        return jsonify({'status': 'unhealthy', 'database': 'disconnected', 'error': str(e), 'timestamp': datetime.now().isoformat()}), 500

# --- Plotting Functions ---
def create_gauge(percentage, title, main_color, bg_color='#404040'):
    fig, ax = plt.subplots(figsize=(3, 3), facecolor='none')
    ax.set_facecolor('none')
    center, radius, start_angle, end_angle = (0.5, 0.5), 0.3, 200, -20
    total_angle = start_angle - end_angle
    ax.add_patch(patches.Wedge(center, radius + 0.08, end_angle, start_angle, width=0.16, facecolor=bg_color, alpha=0.5))
    if percentage > 0:
        filled_angle_span = (percentage / 100) * total_angle
        ax.add_patch(patches.Wedge(center, radius + 0.08, start_angle - filled_angle_span, start_angle, width=0.16, facecolor=main_color, alpha=0.9))
    needle_angle = start_angle - (percentage / 100) * total_angle
    needle_x = center[0] + (radius - 0.02) * np.cos(np.radians(needle_angle))
    needle_y = center[1] + (radius - 0.02) * np.sin(np.radians(needle_angle))
    ax.plot([center[0], needle_x], [center[1], needle_y], color='white', linewidth=3, solid_capstyle='round')
    ax.add_patch(patches.Circle(center, 0.02, color='white', zorder=10))
    ax.text(center[0], center[1] - 0.08, f"{percentage:.1f}%", ha='center', va='center', fontsize=32, fontweight='bold', color='white')
    ax.text(center[0], center[1] - 0.15, title, ha='center', va='center', fontsize=12, fontweight='bold', color='#AAAAAA')
    return fig

# --- Helper function for filtering data ---
def get_filtered_data(args):
    """Applies all filters to the main dataframe based on request arguments."""
    df = get_data()
    if df.empty:
        return df

    filtered_df = df.copy()
    start_date = args.get('start_date')
    end_date = args.get('end_date')
    if start_date: filtered_df = filtered_df[filtered_df['date'] >= pd.to_datetime(start_date)]
    if end_date: filtered_df = filtered_df[filtered_df['date'] <= pd.to_datetime(end_date)]

    filter_map = {'department': 'department', 'class': 'class', 'teacher': 'teacher_name', 'subject': 'subject', 'student': 'name', 'lecture_time': 'lecture_time', 'day': 'day'}
    for key, column in filter_map.items():
        values = args.getlist(key)
        if values:
            filtered_df = filtered_df[filtered_df[column].isin(values)]

    return filtered_df

@app.route('/api/filters')
def get_filters_api():
    """Provides all unique filter options to the frontend."""
    df = get_data()
    if df.empty:
        return jsonify({"departments": [], "classes": [], "teachers": [], "subjects": [], "students": [], "lecture_times": [], "days": []})

    lecture_order = ["09:00-10:00", "10:00-11:00", "11:00-12:00", "12:00-01:00", "01:30-02:30", "02:30-03:30"]
    available_times = df["lecture_time"].dropna().str.strip().unique()
    weekday_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    available_days = df["day"].dropna().unique()

    filters = {
        "departments": sorted(df["department"].unique().tolist()),
        "classes": sorted(df["class"].unique().tolist()),
        "teachers": sorted(df["teacher_name"].unique().tolist()),
        "subjects": sorted(df["subject"].unique().tolist()),
        "students": sorted(df["name"].unique().tolist()),
        "lecture_times": [time for time in lecture_order if time in available_times],
        "days": [day for day in weekday_order if day in available_days]
    }
    return jsonify(filters)

@app.route('/plot/<plot_name>')
def plot_generator(plot_name):
    filtered_df = get_filtered_data(request.args)
    fig = None
    color_palette = sns.color_palette("viridis", 15)

    if filtered_df.empty:
        fig, ax = plt.subplots(figsize=(6, 4), facecolor='none')
        ax.set_facecolor('none')
        ax.text(0.5, 0.5, 'No data for selected filters', ha='center', va='center', fontsize=12, color='white')
        ax.axis('off')
    else:
        plt.rcParams.update({'text.color': 'white', 'axes.labelcolor': 'white', 'xtick.color': 'white', 'ytick.color': 'white'})

        if 'gauge' in plot_name:
            total = len(filtered_df)
            present = len(filtered_df[filtered_df["attendance"].str.upper() == "P"])
            attendance_pct = round((present / total) * 100, 2) if total > 0 else 0
            fig = create_gauge(attendance_pct, "P R E S E N T", "#00FF7F") if plot_name == 'present_gauge' else create_gauge(100 - attendance_pct, "A B S E N T", "#FF6347")

        elif plot_name.startswith('absentee_by'):
            filtered_df["is_absent"] = filtered_df["attendance"].str.upper() == "A"
            fig, ax = plt.subplots(figsize=(8, 5), facecolor='none')
            ax.set_facecolor('none')
            ax.grid(False) # Remove grid lines
            hue_col = "department" if plot_name == 'absentee_by_dept' else "class"
            title = f"Absent % by Student ({hue_col.title()})"
            trend_df = filtered_df.groupby(["name", hue_col]).agg(absent=('is_absent', 'sum'), total=('attendance', 'count')).reset_index()
            if not trend_df.empty and trend_df['total'].sum() > 0:
                trend_df["absent_pct"] = (trend_df["absent"] / trend_df["total"] * 100).round(2)
                sns.lineplot(data=trend_df.sort_values("absent_pct", ascending=False), x="name", y="absent_pct", hue=hue_col, marker="o", ax=ax, palette=color_palette)
            ax.set_title(title)
            ax.tick_params(axis='x', rotation=45)
            if ax.get_legend():
                ax.get_legend().get_title().set_color('white')
                for text in ax.get_legend().get_texts(): text.set_color('white')

        elif plot_name.startswith('donut_by'):
            fig, ax = plt.subplots(figsize=(7, 7), facecolor='none')
            ax.set_facecolor('none')
            column = "class" if plot_name == 'donut_by_class' else "lecture_time"
            title = f"Attendance % by {column.replace('_', ' ').title()}"
            counts = filtered_df[filtered_df["attendance"].str.upper() == "P"][column].value_counts()
            if column == 'lecture_time':
                lecture_order = ["09:00-10:00", "10:00-11:00", "11:00-12:00", "12:00-01:00", "01:30-02:30", "02:30-03:30"]
                counts = counts.reindex(lecture_order).dropna()

            wedges, texts, autotexts = ax.pie(counts, labels=counts.index, autopct='%1.1f%%', startangle=90, colors=color_palette, pctdistance=0.85, wedgeprops=dict(width=0.4, edgecolor='k'))
            plt.setp(autotexts, size=10, weight="bold", color="white")
            plt.setp(texts, size=12, color="white")
            ax.set_title(title)

        elif plot_name.startswith('bar_by_day'):
            fig, ax = plt.subplots(figsize=(8, 5), facecolor='none')
            ax.set_facecolor('none')
            ax.grid(False) # Remove grid lines
            hue_col = "class" if plot_name == 'bar_by_day_class' else "department"
            title = f"Attendance % by Day and {hue_col.title()}"
            pct_df = filtered_df.groupby(["day", hue_col]).agg(total=('attendance', 'count'), present=('attendance', lambda x: x.str.upper().eq('P').sum())).reset_index()
            if not pct_df.empty and pct_df['total'].sum() > 0:
                pct_df["attendance_pct"] = ((pct_df["present"] / pct_df["total"]) * 100).round(2)
                weekday_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
                sns.barplot(data=pct_df, x="day", y="attendance_pct", hue=hue_col, ax=ax, order=weekday_order, palette=color_palette)
            ax.set_title(title)
            ax.set_ylim(0, 100)
            if ax.get_legend(): ax.get_legend().get_title().set_color('white')

    buf = io.BytesIO()
    fig.savefig(buf, format='png', transparent=True)
    plt.close(fig)
    buf.seek(0)
    return send_file(buf, mimetype='image/png')

@app.route('/api/table/<table_name>')
def table_generator(table_name):
    filtered_df = get_filtered_data(request.args)
    if filtered_df.empty:
        return jsonify([])

    if table_name == 'top_defaulters':
        filtered_df["is_absent"] = filtered_df["attendance"].str.upper() == "A"
        summary = filtered_df.groupby(["name", "department", "class"]).agg(total_attendance=('attendance', 'count'), total_absent=('is_absent', 'sum')).reset_index()
        if not summary.empty and summary['total_attendance'].sum() > 0:
            summary["absent_pct"] = round((summary["total_absent"] / summary["total_attendance"]) * 100, 2)
            summary = summary.sort_values("absent_pct", ascending=False).head(10)
            summary["absent_pct_formatted"] = summary["absent_pct"].apply(
                lambda val: f'<span style="color:{"red" if val > 80 else "orange" if val > 50 else "lightgreen"}; font-weight:bold">{"🔺" if val > 50 else "🔻"} {val:.2f}%</span>'
            )
            summary = summary.rename(columns={"name": "Name", "department": "Department", "class": "Class", "total_attendance": "Total Lectures", "total_absent": "Total Absent", "absent_pct_formatted": "Absent %"})
            return jsonify(summary[["Name", "Department", "Class", "Total Lectures", "Total Absent", "Absent %"]].to_dict(orient='records'))

    return jsonify([])

if __name__ == '__main__':
    app.run(debug=True)
