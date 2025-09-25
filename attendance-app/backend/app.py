import csv
import os
from io import StringIO
from flask import Flask, Response, jsonify, request, session, redirect, url_for, render_template, send_from_directory
from flask_cors import CORS
from werkzeug.utils import secure_filename
import logging
from datetime import datetime, timedelta
from openpyxl import load_workbook
import psycopg2
from functools import wraps
from werkzeug.security import generate_password_hash

app = Flask(__name__)
CORS(app)

app.secret_key = 'your-secret-key-change-this-in-production'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SECURE'] = False
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)
app.config['UPLOAD_FOLDER'] = 'uploads/'
app.config['ALLOWED_EXTENSIONS'] = {'xlsx', 'xls', 'csv'}
app.config['PG_HOST'] = 'aws-0-ap-south-1.pooler.supabase.com'
app.config['PG_USER'] = 'postgres.avqpzwgdylnklbkyqukp'
app.config['PG_PASSWORD'] = 'asBjLmDfKfoZPVt9'
app.config['PG_DB'] = 'postgres'
app.config['sslmode']='require'

logging.basicConfig(level=logging.INFO)

def get_pg_connection():
    return psycopg2.connect(
        host=app.config['PG_HOST'],
        database=app.config['PG_DB'],
        user=app.config['PG_USER'],
        password=app.config['PG_PASSWORD'],
        port=6543,
        sslmode=app.config['sslmode']
    )

def get_user_from_db(email):
    try:
        connection = get_pg_connection()
        cursor = connection.cursor()
        cursor.execute("""
            SELECT id, email, password_hash, created_at, role
            FROM users 
            WHERE email = %s
        """, (email,))
        user_row = cursor.fetchone()
        cursor.close()
        connection.close()
        if user_row:
            return {
                'id': user_row[0],
                'email': user_row[1],
                'password_hash': user_row[2],
                'created_at': user_row[3],
                'role': user_row[4] if user_row[4] else 'user',
                'name': user_row[1].split('@')[0].replace('.', ' ').title(),
            }
        return None
    except Exception as e:
        app.logger.error(f"Database error in get_user_from_db: {e}")
        return None

def verify_password_crypt(password, password_hash):
    try:
        connection = get_pg_connection()
        cursor = connection.cursor()
        cursor.execute("""
            SELECT crypt(%s, %s) = %s AS password_match
        """, (password, password_hash, password_hash))
        result = cursor.fetchone()
        cursor.close()
        connection.close()
        return result[0] if result else False
    except Exception as e:
        app.logger.error(f"Password verification error: {e}")
        return False

def create_user_with_crypt(email, password, role='user'):
    try:
        connection = get_pg_connection()
        cursor = connection.cursor()
        cursor.execute("""
            INSERT INTO users (email, password_hash, role)
            VALUES (%s, crypt(%s, gen_salt('bf')), %s)
            ON CONFLICT (email) DO UPDATE 
            SET password_hash = crypt(EXCLUDED.password_hash, gen_salt('bf')), 
                role = EXCLUDED.role
        """, (email, password, role))
        connection.commit()
        cursor.close()
        connection.close()
        return True
    except Exception as e:
        app.logger.error(f"Error creating user: {e}")
        return False

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.is_json or request.headers.get('Content-Type') == 'application/json' or 'api/' in request.path:
                return jsonify({'error': 'Authentication required', 'redirect': '/login'}), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.is_json or request.headers.get('Content-Type') == 'application/json' or 'api/' in request.path:
                return jsonify({'error': 'Authentication required', 'redirect': '/login'}), 401
            return redirect(url_for('login_page'))
        
        user_role = session.get('user_role', 'user')
        if user_role != 'admin':
            if request.is_json or request.headers.get('Content-Type') == 'application/json' or 'api/' in request.path:
                return jsonify({
                    'error': 'Admin access required', 
                    'message': 'You do not have permission to access this resource',
                    'redirect': '/'
                }), 403
            return redirect(url_for('home'))
        
        return f(*args, **kwargs)
    return decorated_function

@app.route('/login')
def login_page():
    if 'user_id' in session:
        return redirect(url_for('home'))
    return render_template('login.html')

@app.route("/")
def root():
    if 'user_id' not in session:
        return redirect(url_for('login_page'))
    return redirect(url_for('home'))

@app.route("/home")
@login_required
def home():
    return render_template("index.html")

@app.route("/attendance")
@login_required
def attendance_tracker():
    return render_template("Attendance.html")

@app.route("/admin-students")
@admin_required
def admin_student():
    return render_template("admin-students.html")

@app.route("/admin-teacher")
@admin_required
def admin_teacher():
    return render_template("admin-teacher.html")

@app.route("/dashboard")
@login_required
def dashboard_page():
    return render_template("dashboard.html")

@app.route('/static/<path:filename>')
def serve_static(filename):
    return send_from_directory('static', filename)

@app.route('/api/login', methods=['POST'])
def api_login():
    try:
        data = request.get_json()
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        remember = data.get('remember', False)
        user = get_user_from_db(email)
        
        if user and verify_password_crypt(password, user['password_hash']):
            session.clear()
            session['user_id'] = user['email']
            session['user_name'] = user['name']
            session['user_role'] = user['role']
            session['user_db_id'] = user['id']
            session.permanent = remember
            
            return jsonify({
                'success': True,
                'message': 'Login successful',
                'redirect': '/',
                'user': {
                    'email': user['email'],
                    'name': user['name'],
                    'role': user['role']
                }
            })
        else:
            return jsonify({'success': False, 'message': 'Invalid email or password'}), 401
    except Exception as e:
        return jsonify({'success': False, 'message': 'Login failed. Please try again.'}), 500

@app.route('/api/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({'success': True, 'message': 'Logged out successfully', 'redirect': '/login'})

@app.route('/api/forgot-password', methods=['POST'])
def api_forgot_password():
    try:
        data = request.get_json()
        email = data.get('email', '').strip().lower()
        if not email:
            return jsonify({'success': False, 'message': 'Email is required'}), 400
        user = get_user_from_db(email)
        if not user:
            return jsonify({'success': False, 'message': 'Email not found'}), 404
        return jsonify({'success': True, 'message': 'Password reset link sent to your email'})
    except Exception as e:
        return jsonify({'success': False, 'message': 'Failed to send reset link'}), 500

@app.route('/api/register', methods=['POST'])
def api_register():
    try:
        data = request.get_json()
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        role = data.get('role', 'user')
        if role not in ['admin', 'user']:
            role = 'user'
        if not email or not password:
            return jsonify({'success': False, 'message': 'Email and password are required'}), 400
        if len(password) < 6:
            return jsonify({'success': False, 'message': 'Password must be at least 6 characters long'}), 400
        existing_user = get_user_from_db(email)
        if existing_user:
            return jsonify({'success': False, 'message': 'User already exists'}), 409
        if create_user_with_crypt(email, password, role):
            return jsonify({'success': True, 'message': f'User {email} created successfully with role {role}'})
        else:
            return jsonify({'success': False, 'message': 'Registration failed. Please try again.'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': 'Registration failed. Please try again.'}), 500

@app.route('/api/user-info', methods=['GET'])
@login_required
def api_user_info():
    return jsonify({
        'user_id': session.get('user_id'),
        'user_name': session.get('user_name'),
        'user_role': session.get('user_role'),
        'user_db_id': session.get('user_db_id')
    })

@app.route('/departments', methods=['GET'])
@login_required
def get_departments():
    connection = get_pg_connection()
    cursor = connection.cursor()
    cursor.execute("SELECT DISTINCT department FROM Teachers ORDER BY department ASC")
    departments = cursor.fetchall()
    cursor.close()
    connection.close()
    return jsonify([dept[0] for dept in departments])

@app.route('/teachers', methods=['GET'])
@login_required
def get_teachers():
    department = request.args.get('department')
    connection = get_pg_connection()
    cursor = connection.cursor()
    cursor.execute("SELECT DISTINCT teacher_name FROM Teachers WHERE department = %s ORDER BY teacher_name ASC", (department,))
    teachers = cursor.fetchall()
    cursor.close()
    connection.close()
    return jsonify([teacher[0] for teacher in teachers])

@app.route('/student-classes', methods=['GET'])
@login_required
def get_student_classes():
    connection = get_pg_connection()
    cursor = connection.cursor()
    cursor.execute("SELECT DISTINCT class FROM Students ORDER BY class ASC")
    classes = cursor.fetchall()
    cursor.close()
    connection.close()
    return jsonify([class_[0] for class_ in classes])

@app.route('/subjects', methods=['GET'])
@login_required
def get_subjects():
    department = request.args.get('department')
    class_name = request.args.get('class')
    teacher_name = request.args.get('teacher_name')
    if not all([department, class_name, teacher_name]):
        return jsonify({"error": "Parameters 'department', 'class', and 'teacher_name' are required"}), 400
    connection = get_pg_connection()
    cursor = connection.cursor()
    query = "SELECT DISTINCT subject FROM Teachers WHERE department = %s AND class = %s AND teacher_name = %s ORDER BY subject ASC"
    cursor.execute(query, (department, class_name, teacher_name))
    subjects = cursor.fetchall()
    cursor.close()
    connection.close()
    return jsonify([subject[0] for subject in subjects])

import logging # Make sure this is at the top of your file
from datetime import datetime # And this too

import logging # Make sure this is at the top of your file

@app.route('/time-slots', methods=['GET'])
@login_required
def get_time_slots():
    connection = get_pg_connection()
    cursor = connection.cursor()
    
    cursor.execute("SELECT DISTINCT time_slot FROM Teachers WHERE time_slot IS NOT NULL AND time_slot != ''")
    time_slots_raw = cursor.fetchall()
    cursor.close()
    connection.close()

    time_slots_list = [item[0] for item in time_slots_raw]
    
    # This new sort key understands PM times
    def sort_key(time_slot_str):
        try:
            start_time_str = time_slot_str.split('-')[0].strip()
            hour, minute = map(int, start_time_str.split(':'))
            
            # **THE FIX:** If the hour is less than 8, treat it as PM by adding 12.
            # For example, 01:30 becomes 13:30 for sorting.
            if hour < 8:
                hour += 12
            
            # Return a tuple for sorting. (9, 0) comes before (13, 30).
            return (hour, minute)

        except (ValueError, IndexError):
            logging.warning(f"Could not parse time slot: '{time_slot_str}'. Placing it at the end.")
            # Return a tuple that will always be sorted last
            return (23, 59)

    sorted_time_slots = sorted(time_slots_list, key=sort_key)
    
    return jsonify(sorted_time_slots)    

@app.route('/students', methods=['GET'])
@login_required
def get_students():
    query_param = request.args.get('query', '').strip()
    department = request.args.get('department', '').strip()
    student_class = request.args.get('class', '').strip()
    connection = get_pg_connection()
    cursor = connection.cursor()
    if query_param:
        cursor.execute("SELECT roll_number, name FROM Students WHERE roll_number LIKE %s OR name LIKE %s", (f"%{query_param}%", f"%{query_param}%"))
    elif department and student_class:
        cursor.execute("SELECT roll_number, name FROM Students WHERE department = %s AND class = %s", (department, student_class))
    elif department:
        cursor.execute("SELECT roll_number, name FROM Students WHERE department = %s", (department,))
    elif student_class:
        cursor.execute("SELECT roll_number, name FROM Students WHERE class = %s", (student_class,))
    else:
        cursor.execute("SELECT roll_number, name FROM Students")
    students = cursor.fetchall()
    cursor.close()
    connection.close()
    return jsonify([{"roll_number": student[0], "name": student[1]} for student in students])

@app.route('/attendance', methods=['POST'])
@login_required
def save_attendance():
    data = request.json
    connection = get_pg_connection()
    cursor = connection.cursor()
    for record in data['attendance_records']:
        cursor.execute(""" 
            INSERT INTO Attendance (date, roll_number, name, department, class, subject, teacher_name, lecture_time, attendance)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (date, lecture_time, roll_number) 
            DO UPDATE SET attendance = EXCLUDED.attendance
        """, (
            record['date'], 
            record['roll_number'], 
            record['name'], 
            record['department'], 
            record['class'], 
            record['subject'], 
            record['teacher_name'], 
            record['lecture_time'], 
            record['attendance']
        ))
    connection.commit()
    cursor.close()
    connection.close()
    return jsonify({"status": "success"})

@app.route('/attendance-data', methods=['GET'])
@login_required
def get_attendance_data():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    department = request.args.get('department', '').strip()
    class_name = request.args.get('class', '').strip()
    if not start_date or not end_date:
        return jsonify({'message': 'Start date and end date are required'}), 400
    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'message': 'Invalid date format. Use YYYY-MM-DD.'}), 400
    try:
        connection = get_pg_connection()
        if connection is None:
            return jsonify({'message': 'Database connection failed'}), 500
        with connection.cursor() as cursor:
            query = "SELECT date, roll_number, name, department, class, attendance, lecture_time FROM Attendance WHERE date BETWEEN %s AND %s"
            filters = [start_date, end_date]
            if department:
                query += " AND department = %s"
                filters.append(department)
            if class_name:
                query += " AND class = %s"
                filters.append(class_name)
            cursor.execute(query, tuple(filters))
            rows = cursor.fetchall()
            if not rows:
                return jsonify([])
            result = [{'date': row[0], 'roll_number': row[1], 'name': row[2], 'department': row[3], 'class': row[4], 'attendance': row[5], 'lecture_time': row[6]} for row in rows]
            connection.close()
            return jsonify(result)
    except Exception as e:
        logging.error(f"Error fetching attendance data: {e}")
        return jsonify({'message': 'An error occurred while fetching attendance data'}), 500

@app.route('/attendance-csv', methods=['GET'])
@login_required
def download_attendance_csv():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    if not start_date or not end_date:
        return jsonify({'message': 'Start and end dates are required'}), 400
    try:
        connection = get_pg_connection()
        cursor = connection.cursor()
        query = "SELECT date, roll_number, name, department, class, attendance, lecture_time FROM Attendance WHERE date BETWEEN %s AND %s"
        cursor.execute(query, (start_date, end_date))
        rows = cursor.fetchall()
        cursor.close()
        connection.close()
        if not rows:
            return jsonify({'message': 'No attendance records found for the given dates'}), 404
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['Date', 'Roll Number', 'Name', 'Department', 'Class', 'Attendance', 'Lecture Time'])
        writer.writerows(rows)
        output.seek(0)
        return Response(output.getvalue(), mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=attendance_data.csv'})
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'message': 'An error occurred while generating the CSV'}), 500

@app.route('/student-departments', methods=['GET'])
@login_required
def get_student_departments():
    connection = get_pg_connection()
    cursor = connection.cursor()
    cursor.execute("SELECT DISTINCT department FROM Students ORDER BY department ASC")
    departments = cursor.fetchall()
    cursor.close()
    connection.close()
    return jsonify([dept[0] for dept in departments])

@app.route('/submit', methods=['POST'])
@admin_required
def submit_student():
    try:
        data = request.json
        roll_number = data.get('roll_number')
        name = data.get('name')
        department = data.get('department')
        class_value = data.get('class')
        if not all([roll_number, name, department, class_value]):
            return jsonify({'success': False, 'message': 'All fields are required'}), 400
        connection = get_pg_connection()
        cursor = connection.cursor()
        cursor.execute("SELECT 1 FROM Students WHERE roll_number = %s", (roll_number,))
        existing_student = cursor.fetchone()
        if existing_student:
            cursor.close()
            connection.close()
            return jsonify({'success': False, 'message': 'Roll number already exists'}), 409
        cursor.execute("""
            INSERT INTO Students (roll_number, name, department, class)
            VALUES (%s, %s, %s, %s)
        """, (roll_number, name, department, class_value))
        connection.commit()
        cursor.close()
        connection.close()
        return jsonify({'success': True, 'message': 'Student added successfully'})
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'success': False, 'message': 'Failed to submit student data'}), 500

def process_excel(file_path):
    data = []
    workbook = load_workbook(file_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return data

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

@app.route('/upload-file', methods=['POST'])
@admin_required
def upload_file():
    try:
        file = request.files.get('file')
        if not file or file.filename == '':
            return jsonify({'message': 'No file selected'}), 400
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            file.save(file_path)
            data = process_excel(file_path)
            total_records = len(data)
            valid_records = 0
            error_records = 0
            valid_data = []
            error_data = []
            connection = get_pg_connection()
            cursor = connection.cursor()
            for row in data:
                try:
                    roll_number, name, department, class_value = row
                    cursor.execute("SELECT 1 FROM Students WHERE roll_number = %s", (roll_number,))
                    existing = cursor.fetchone()
                    if existing:
                        error_data.append((roll_number, name, department, class_value, "Duplicate Roll Number"))
                        error_records += 1
                        continue
                    cursor.execute("""
                        INSERT INTO Students (roll_number, name, department, class)
                        VALUES (%s, %s, %s, %s)
                    """, (roll_number, name, department, class_value))
                    valid_data.append((roll_number, name, department, class_value))
                    valid_records += 1
                except Exception as e:
                    error_data.append(row + (str(e),))
                    error_records += 1
            if valid_data:
                cursor.executemany("INSERT INTO Validrecords (roll_number, name, department, class) VALUES (%s, %s, %s, %s)", valid_data)
            if error_data:
                cursor.executemany("INSERT INTO Errorrecords (roll_number, name, department, class, error_message) VALUES (%s, %s, %s, %s, %s)", error_data)
            upload_time = datetime.now()
            cursor.execute("INSERT INTO UploadHistory (uploader_name, total_records, valid_records, error_records, upload_time) VALUES (%s, %s, %s, %s, %s)", (session.get('user_name', 'Unknown'), total_records, valid_records, error_records, upload_time))
            connection.commit()
            cursor.close()
            connection.close()
            return jsonify({'message': 'File processed successfully.', 'total_records': total_records, 'valid_records': valid_records, 'error_records': error_records}), 200
        else:
            return jsonify({'message': 'Invalid file type.'}), 400
    except Exception as e:
        print(f"Error while uploading file: {e}")
        return jsonify({'message': 'An error occurred while processing the file.'}), 500

@app.route('/upload-history', methods=['GET'])
@admin_required
def upload_history():
    try:
        connection = get_pg_connection()
        cursor = connection.cursor()
        cursor.execute("SELECT uploader_name, total_records, valid_records, error_records, upload_time FROM UploadHistory")
        rows = cursor.fetchall()
        history = [{'uploader_name': row[0], 'total_records': row[1], 'valid_records': row[2], 'error_records': row[3], 'upload_time': row[4]} for row in rows]
        cursor.close()
        connection.close()
        return jsonify(history)
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'message': 'Failed to fetch upload history'}), 500

@app.route('/valid-records-csv', methods=['GET'])
@admin_required
def download_valid_csv():
    try:
        connection = get_pg_connection()
        cursor = connection.cursor()
        cursor.execute("SELECT roll_number, name, department, class FROM Validrecords")
        rows = cursor.fetchall()
        cursor.close()
        connection.close()
        if not rows:
            return jsonify({'message': 'No valid records found'}), 404
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['Date', 'Roll Number', 'Name', 'Department', 'Class', 'Attendance', 'Lecture Time'])
        writer.writerows(rows)
        output.seek(0)
        return Response(output.getvalue(), mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=valid_records.csv'})
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'message': 'An error occurred while generating the CSV'}), 500

@app.route('/upload-history/row-count', methods=['GET'])
@admin_required
def get_row_count():
    try:
        uploader_name = request.args.get('uploader_name', '').strip()
        if not uploader_name:
            return jsonify({'message': 'Uploader name is required'}), 400
        connection = get_pg_connection()
        cursor = connection.cursor()
        cursor.execute("SELECT COUNT(*) FROM UploadHistory WHERE uploader_name = %s", (uploader_name,))
        row_count = cursor.fetchone()[0]
        cursor.close()
        connection.close()
        return jsonify({'uploader_name': uploader_name, 'row_count': row_count})
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'message': 'An error occurred while fetching row count'}), 500

@app.route('/teachers', methods=['POST'])
@admin_required
def save_teacher():
    try:
        data = request.json
        teacher_name = data.get('teacher_name')
        day = data.get('day')
        subject = data.get('subject')
        time_slot = data.get('time_slot')
        department = data.get('department')
        class_value = data.get('class')
        if not all([teacher_name, day, subject, time_slot, department, class_value]):
            return jsonify({'message': 'All fields are required'}), 400
        connection = get_pg_connection()
        cursor = connection.cursor()
        query = "INSERT INTO Teachers (teacher_name, day, subject, time_slot, department, class) VALUES (%s, %s, %s, %s, %s, %s) ON CONFLICT (subject, time_slot, department, class) DO UPDATE SET teacher_name = EXCLUDED.teacher_name"
        cursor.execute(query, (teacher_name, day, subject, time_slot, department, class_value))
        connection.commit()
        cursor.close()
        connection.close()
        return jsonify({'message': 'Teacher data saved or updated successfully'})
    except Exception as e:
        print(f"Error saving teacher: {e}")
        return jsonify({'message': 'Failed to save teacher data'}), 500

@app.route('/upload-teachers', methods=['POST'])
@admin_required
def upload_teachers():
    try:
        file = request.files.get('file')
        if not file or file.filename == '':
            return jsonify({'message': 'No file selected'}), 400
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            file.save(file_path)
            data = process_excel(file_path)
            total_records = len(data)
            valid_records = 0
            error_records = 0
            valid_data = []
            error_data = []
            connection = get_pg_connection()
            cursor = connection.cursor()
            for row in data:
                try:
                    teacher_name, day, subject, time_slot, department, class_value = row
                    cursor.execute("INSERT INTO Teachers (teacher_name, day, subject, time_slot, department, class) VALUES (%s, %s, %s, %s, %s, %s) ON CONFLICT (subject, time_slot, department, class) DO UPDATE SET teacher_name = EXCLUDED.teacher_name", (teacher_name, day, subject, time_slot, department, class_value))
                    valid_data.append((teacher_name, day, subject, time_slot, department, class_value))
                    valid_records += 1
                except Exception as e:
                    error_data.append(row + (str(e),))
                    error_records += 1
            connection.commit()
            if valid_data:
                cursor.executemany("INSERT INTO ValidTeachers (teacher_name, day, subject, time_slot, department, class) VALUES (%s, %s, %s, %s, %s, %s)", valid_data)
            if error_data:
                cursor.executemany("INSERT INTO ErrorTeachers (teacher_name, day, subject, time_slot, department, class, error_message) VALUES (%s, %s, %s, %s, %s, %s, %s)", error_data)
            connection.commit()
            cursor.close()
            connection.close()
            return jsonify({'message': 'File processed successfully.', 'total_records': total_records, 'valid_records': valid_records, 'error_records': error_records}), 200
        else:
            return jsonify({'message': 'Invalid file type. Only .xlsx, .xls, and .csv files are allowed.'}), 400
    except Exception as e:
        print(f"Error while uploading file: {e}")
        return jsonify({'message': 'An error occurred while processing the file.'}), 500

# Admin utility endpoint to create users - ADMIN ONLY
@app.route('/api/admin/create-user', methods=['POST'])
@admin_required
def admin_create_user():
    try:
        data = request.get_json()
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        role = data.get('role', 'user')
        if role not in ['admin', 'user']:
            role = 'user'
        if not email or not password:
            return jsonify({'success': False, 'message': 'Email and password are required'}), 400
        if create_user_with_crypt(email, password, role):
            return jsonify({'success': True, 'message': f'User {email} created successfully with role {role}'})
        else:
            return jsonify({'success': False, 'message': 'Failed to create user'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': 'Failed to create user'}), 500

@app.route('/api/user-permissions', methods=['GET'])
@login_required
def api_user_permissions():
    user_role = session.get('user_role', 'user')
    permissions = {
        'can_access_admin_students': user_role == 'admin',
        'can_access_admin_teachers': user_role == 'admin',
        'can_upload_files': user_role == 'admin',
        'can_manage_students': user_role == 'admin',
        'can_manage_teachers': user_role == 'admin',
        'can_view_upload_history': user_role == 'admin',
        'user_role': user_role
    }
    return jsonify(permissions)

@app.route('/debug/session')
def debug_session():
    return jsonify({'session_data': dict(session), 'user_id_in_session': 'user_id' in session, 'session_keys': list(session.keys()), 'session_permanent': session.permanent})

@app.route('/test/session')
def test_session():
    session['test'] = 'working'
    return jsonify({'message': 'Session test set'})

@app.route('/test/session/check')
def check_session():
    return jsonify({'test_value': session.get('test', 'not found')})

@app.route('/health', methods=['GET'])
def health_check():
    try:
        # Test database connection
        connection = get_pg_connection()
        cursor = connection.cursor()
        cursor.execute("SELECT 1")
        cursor.close()
        connection.close()
        return jsonify({'status': 'healthy', 'database': 'connected', 'timestamp': datetime.now().isoformat()})
    except Exception as e:
        return jsonify({'status': 'unhealthy', 'database': 'disconnected', 'error': str(e), 'timestamp': datetime.now().isoformat()}), 500

if __name__ == '__main__':
    app.run(debug=True)
